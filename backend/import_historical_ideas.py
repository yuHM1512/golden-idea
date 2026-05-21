from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from sqlalchemy import text

EXTRA_SITE_PACKAGES = os.environ.get("PROJECT_SITE_PACKAGES", "")
for raw_path in EXTRA_SITE_PACKAGES.split(os.pathsep):
    extra_path = raw_path.strip()
    if extra_path and extra_path not in sys.path:
        sys.path.append(extra_path)

from app.database import Base, SessionLocal, engine
from app.config import settings
from app.models.actual_benefit import ActualBenefitEvaluation
from app.models.attachment import FileAttachment
from app.models.idea import Idea, IdeaCategory, IdeaStatus
from app.models.review import IdeaReview, ReviewAction, ReviewLevel
from app.models.score import IdeaScore, K1Type, K2Type, K3MeasureType
from app.models.unit import Unit
from app.models.user import User, UserRole
from app.seed import (
    migrate_file_attachments_drive_columns,
    migrate_idea_bo_phan_column,
    migrate_idea_category_column,
    migrate_idea_participants_column,
    migrate_idea_title_column,
    migrate_ie_review_logic_columns,
    migrate_k3_cost_saved_criteria_codes,
    migrate_payment_slip_amount_default,
    migrate_payment_slip_reward_columns,
    migrate_reward_batch_special_coefficients_column,
    migrate_score_criteria_tables,
    migrate_score_k2_type_column,
    migrate_standardized_idea_replications_table,
    migrate_user_role_column,
    migrate_users_unit_nullable,
    normalize_employee_codes,
    normalize_sample_idea_categories,
    seed_admin_user,
    seed_score_criteria,
    seed_units,
)

APPROVED_STANDARDIZATION = "APPROVED_STANDARDIZATION"
APPROVED_NO_STANDARDIZATION = "APPROVED_NO_STANDARDIZATION"
DEFAULT_XLSX = Path(settings.HISTORICAL_IDEAS_XLSX)

DATA_COL = {
    "submitted_at": 1,
    "stt": 2,
    "full_name": 3,
    "unit": 4,
    "bo_phan": 5,
    "phone": 6,
    "category_text": 7,
    "description": 8,
    "employee_code": 9,
    "film_1": 10,
    "film_2": 11,
    "film_3": 12,
    "film_4": 13,
    "film_5": 14,
    "film_6": 15,
    "total_score_xn_text": 16,
    "reward_text": 17,
    "status_text": 18,
    "email": 19,
    "standardization": 20,
    "position": 21,
    "product_code": 22,
    "year": 23,
    "month": 24,
    "title": 25,
}

SCORE_COL = {
    "id": 1,
    "sub_id": 2,
    "stt": 3,
    "full_name": 4,
    "employee_code": 5,
    "unit": 6,
    "bo_phan": 7,
    "product_code": 8,
    "before_seconds": 9,
    "after_seconds": 10,
    "improvement_ratio": 11,
    "quantity": 12,
    "saved_cost_value": 13,
    "benefit_value": 14,
    "labor_second_price": 15,
    "unmeasurable_flag": 16,
    "total_score_xn": 40,
    "approved_score": 41,
    "note": 42,
    "registered_at": 44,
}

K1_SCORES = {"A1": 10, "A2": 5, "A3": 2}
K2_EASY_SCORES = {"B1": 3, "B2": 3, "B3": 3}
K2_HARD_SCORES = {"B4": 2, "B5": 2, "B6": 2}
K3_TIME_SCORES = {"C1": 60, "C2": 40, "C3": 20, "C4": 10, "C5": 5}
K3_COST_SCORES = {"C6": 100, "C7": 60, "C8": 40, "C9": 20, "C10": 10}
K3_UNMEASURABLE_SCORES = {"C11": 10, "C12": 10, "C13": 10, "C14": 10}

UNIT_ALIASES = {
    "DUYTRUNG": "XNDT",
    "VESTON1": "XN1-V1",
    "VESTON 1": "XN1-V1",
    "XN1": "XN1-V1",
    "VESTON2": "XNV2",
    "XN4": "XN4",
    "CƠĐIỆN": "P. KTCĐ & ĐTMT",
    "P.KD": "P. KDXNK",
    "PKD": "P. KDXNK",
    "QLCL": "P. QLCL",
    "P.QLCL": "P. QLCL",
    "P.TỔNGHỢP": "P. Tổng hợp",
    "PTỔNGHỢP": "P. Tổng hợp",
    "BANTHIẾTBỊ": "Ban Thiết bị",
}


@dataclass
class CandidateScore:
    k1_type: str
    k1_score: int
    k2_type: str
    k2_codes: list[str]
    k2_score: int
    k3_measure_type: str
    k3_option_code: str | None
    k3_codes: list[str]
    k3_score: int
    total_score: int
    diff: int


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_code(value: Any) -> str:
    return str(value or "").strip().upper()


def normalize_unit_key(value: Any) -> str:
    text = normalize_text(value) or ""
    return re.sub(r"[\s._-]+", "", text.upper())


def normalize_status_text(value: Any) -> str:
    return (normalize_text(value) or "").casefold()


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    return None


def parse_numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    numeric = parse_numeric(value)
    if numeric is None:
        return None
    if numeric.is_integer():
        return int(numeric)
    return int(round(numeric))


def parse_drive_file_id(url: str | None) -> str | None:
    if not url:
        return None
    parsed = urlparse(url)
    query_id = parse_qs(parsed.query).get("id")
    if query_id:
        return query_id[0]
    match = re.search(r"/d/([^/]+)", parsed.path)
    if match:
        return match.group(1)
    return None


def fallback_title(description: str | None, stt: Any) -> str:
    text = normalize_text(description) or ""
    if not text:
        return f"Idea #{stt}"
    lines = [line.strip(" -:\t") for line in text.splitlines() if line.strip()]
    if not lines:
        return f"Idea #{stt}"
    first_line = lines[0]
    first_line = re.sub(r"^(trước|sau)\s*(cải tiến|ct|sản xuất)\s*:?\s*", "", first_line, flags=re.IGNORECASE)
    if len(first_line) <= 120:
        return first_line
    return first_line[:117].rstrip() + "..."


def resolve_category(category_text: str | None, description: str | None) -> str:
    text = " ".join(filter(None, [category_text, description])).casefold()
    if "số hoá" in text or "so hoa" in text or "google form" in text or "google sheet" in text or "webapp" in text:
        return IdeaCategory.DIGITIZATION.value
    if any(keyword in text for keyword in ["công cụ", "cu gá", "cử gá", "rập", "form", "thiết bị", "phụ trợ"]):
        return IdeaCategory.TOOLS.value
    if any(keyword in text for keyword in ["qui trình", "quy trình", "phương pháp", "thao tác", "quy cach", "quy cách"]):
        return IdeaCategory.PROCESS.value
    return IdeaCategory.OTHER.value


def initialize_database() -> None:
    import app.models  # noqa: F401

    Base.metadata.create_all(bind=engine)
    migrate_user_role_column()
    migrate_users_unit_nullable()
    migrate_idea_participants_column()
    migrate_idea_bo_phan_column()
    migrate_idea_title_column()
    normalize_employee_codes()
    migrate_idea_category_column()
    migrate_score_k2_type_column()
    migrate_score_criteria_tables()
    migrate_payment_slip_reward_columns()
    migrate_payment_slip_amount_default()
    migrate_reward_batch_special_coefficients_column()
    migrate_ie_review_logic_columns()
    migrate_file_attachments_drive_columns()
    migrate_standardized_idea_replications_table()
    migrate_k3_cost_saved_criteria_codes()
    normalize_sample_idea_categories()
    seed_units()
    seed_score_criteria()
    seed_admin_user()


def _remove_autofilter_from_workbook(source_path: Path) -> Path:
    target_path = Path(tempfile.gettempdir()) / "golden_idea_import_no_autofilter.xlsx"
    with ZipFile(source_path, "r") as zin, ZipFile(target_path, "w", ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                data = re.sub(rb"<autoFilter\b.*?</autoFilter>", b"", data, flags=re.DOTALL)
            zout.writestr(item, data)
    return target_path


def load_workbook_sheets(file_path: Path):
    try:
        workbook = load_workbook(file_path, data_only=False)
    except ValueError as exc:
        error_text = f"{exc!r} {exc.__cause__!r}"
        if "customFilter" not in error_text and "Value must be either numerical" not in error_text:
            raise
        cleaned_path = _remove_autofilter_from_workbook(file_path)
        workbook = load_workbook(cleaned_path, data_only=False)
        print(f"Workbook autoFilter XML was stripped for openpyxl compatibility: {cleaned_path}")

    score_sheet_name = next(name for name in workbook.sheetnames if name.strip().lower() == "điểm")
    return workbook, workbook["data"], workbook[score_sheet_name]


def read_data_rows(data_ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, data_ws.max_row + 1):
        stt = data_ws.cell(row_idx, DATA_COL["stt"]).value
        if stt in (None, ""):
            continue
        attachments: list[dict[str, str]] = []
        for offset, col_name in enumerate(["film_1", "film_2", "film_3", "film_4", "film_5", "film_6"], start=1):
            cell = data_ws.cell(row_idx, DATA_COL[col_name])
            url = cell.hyperlink.target if cell.hyperlink else normalize_text(cell.value)
            if not url:
                continue
            attachments.append({"index": offset, "url": url})
        rows.append(
            {
                "row_idx": row_idx,
                "stt": stt,
                "submitted_at": parse_datetime(data_ws.cell(row_idx, DATA_COL["submitted_at"]).value),
                "full_name": normalize_text(data_ws.cell(row_idx, DATA_COL["full_name"]).value),
                "unit": normalize_text(data_ws.cell(row_idx, DATA_COL["unit"]).value),
                "bo_phan": normalize_text(data_ws.cell(row_idx, DATA_COL["bo_phan"]).value),
                "phone": normalize_text(data_ws.cell(row_idx, DATA_COL["phone"]).value),
                "category_text": normalize_text(data_ws.cell(row_idx, DATA_COL["category_text"]).value),
                "description": normalize_text(data_ws.cell(row_idx, DATA_COL["description"]).value),
                "employee_code": normalize_code(data_ws.cell(row_idx, DATA_COL["employee_code"]).value),
                "status_text": normalize_text(data_ws.cell(row_idx, DATA_COL["status_text"]).value),
                "email": normalize_text(data_ws.cell(row_idx, DATA_COL["email"]).value),
                "standardization": normalize_text(data_ws.cell(row_idx, DATA_COL["standardization"]).value),
                "position": normalize_text(data_ws.cell(row_idx, DATA_COL["position"]).value),
                "product_code": normalize_text(data_ws.cell(row_idx, DATA_COL["product_code"]).value),
                "title": normalize_text(data_ws.cell(row_idx, DATA_COL["title"]).value),
                "attachments": attachments,
            }
        )
    return rows


def read_score_rows(score_ws) -> dict[Any, dict[str, Any]]:
    chosen: dict[Any, dict[str, Any]] = {}
    for row_idx in range(2, score_ws.max_row + 1):
        stt = score_ws.cell(row_idx, SCORE_COL["stt"]).value
        if stt in (None, "", "NA"):
            continue
        total_score = parse_int(score_ws.cell(row_idx, SCORE_COL["total_score_xn"]).value)
        if total_score is None or total_score <= 0:
            continue
        criteria_values = {
            code: parse_numeric(score_ws.cell(row_idx, excel_col).value)
            for code, excel_col in {
                **{code: 16 + idx for idx, code in enumerate(["A1", "A2", "A3"], start=1)},
            }.items()
        }
        criteria_values.update({code: parse_numeric(score_ws.cell(row_idx, col).value) for code, col in {
            "B1": 20, "B2": 21, "B3": 22, "B4": 23, "B5": 24, "B6": 25,
            "C1": 26, "C2": 27, "C3": 28, "C4": 29, "C5": 30, "C6": 31, "C7": 32,
            "C8": 33, "C9": 34, "C10": 35, "C11": 36, "C12": 37, "C13": 38, "C14": 39,
        }.items()})
        row = {
            "row_idx": row_idx,
            "stt": stt,
            "full_name": normalize_text(score_ws.cell(row_idx, SCORE_COL["full_name"]).value),
            "employee_code": normalize_code(score_ws.cell(row_idx, SCORE_COL["employee_code"]).value),
            "before_seconds": parse_numeric(score_ws.cell(row_idx, SCORE_COL["before_seconds"]).value),
            "after_seconds": parse_numeric(score_ws.cell(row_idx, SCORE_COL["after_seconds"]).value),
            "quantity": parse_int(score_ws.cell(row_idx, SCORE_COL["quantity"]).value),
            "benefit_value": parse_numeric(score_ws.cell(row_idx, SCORE_COL["benefit_value"]).value),
            "labor_second_price": parse_numeric(score_ws.cell(row_idx, SCORE_COL["labor_second_price"]).value) or 6.14,
            "total_score_xn": total_score,
            "approved_score": parse_int(score_ws.cell(row_idx, SCORE_COL["approved_score"]).value),
            "note": normalize_text(score_ws.cell(row_idx, SCORE_COL["note"]).value),
            "registered_at": parse_datetime(score_ws.cell(row_idx, SCORE_COL["registered_at"]).value),
            "criteria": criteria_values,
        }
        existing = chosen.get(stt)
        if existing is None:
            chosen[stt] = row
            continue
        existing_priority = (1 if existing["approved_score"] and existing["approved_score"] > 0 else 0, existing["row_idx"])
        current_priority = (1 if row["approved_score"] and row["approved_score"] > 0 else 0, row["row_idx"])
        if current_priority >= existing_priority:
            chosen[stt] = row
    return chosen


def choose_detailed_score(score_row: dict[str, Any]) -> CandidateScore | None:
    positive_k1 = [code for code in K1_SCORES if (score_row["criteria"].get(code) or 0) > 0]
    if not positive_k1:
        return None

    k2_candidates: list[tuple[str, list[str], int]] = []
    easy_codes = [code for code in K2_EASY_SCORES if (score_row["criteria"].get(code) or 0) > 0]
    hard_codes = [code for code in K2_HARD_SCORES if (score_row["criteria"].get(code) or 0) > 0]
    if easy_codes:
        k2_candidates.append((K2Type.EASY.value, easy_codes, sum(K2_EASY_SCORES[code] for code in easy_codes)))
    if hard_codes:
        k2_candidates.append((K2Type.HARD.value, hard_codes, sum(K2_HARD_SCORES[code] for code in hard_codes)))
    if not k2_candidates:
        k2_candidates.append((K2Type.EASY.value, [], 0))

    k3_candidates: list[tuple[str, str | None, list[str], int]] = []
    for code, score in K3_TIME_SCORES.items():
        if (score_row["criteria"].get(code) or 0) > 0:
            k3_candidates.append((K3MeasureType.TIME_SAVED.value, code, [], score))
    for code, score in K3_COST_SCORES.items():
        if (score_row["criteria"].get(code) or 0) > 0:
            k3_candidates.append((K3MeasureType.COST_SAVED.value, code, [], score))
    unmeasurable_codes = [code for code in K3_UNMEASURABLE_SCORES if (score_row["criteria"].get(code) or 0) > 0]
    if unmeasurable_codes:
        k3_candidates.append(
            (
                K3MeasureType.UNMEASURABLE.value,
                None,
                unmeasurable_codes,
                sum(K3_UNMEASURABLE_SCORES[code] for code in unmeasurable_codes),
            )
        )
    if not k3_candidates:
        return None

    target_score = score_row["total_score_xn"]
    best: CandidateScore | None = None
    for k1_code in positive_k1:
        for k2_type, k2_codes, k2_score in k2_candidates:
            for k3_type, k3_option_code, k3_codes, k3_score in k3_candidates:
                total_score = K1_SCORES[k1_code] + k2_score + k3_score
                candidate = CandidateScore(
                    k1_type=k1_code,
                    k1_score=K1_SCORES[k1_code],
                    k2_type=k2_type,
                    k2_codes=k2_codes,
                    k2_score=k2_score,
                    k3_measure_type=k3_type,
                    k3_option_code=k3_option_code,
                    k3_codes=k3_codes,
                    k3_score=k3_score,
                    total_score=total_score,
                    diff=abs(total_score - target_score),
                )
                if best is None:
                    best = candidate
                    continue
                if candidate.diff < best.diff:
                    best = candidate
                    continue
                if candidate.diff == best.diff and candidate.total_score <= best.total_score:
                    best = candidate
    return best


def resolve_standardization_result(data_row: dict[str, Any], score_row: dict[str, Any] | None) -> str | None:
    if not score_row or not score_row.get("approved_score") or score_row["approved_score"] <= 0:
        return None
    text = normalize_status_text(data_row.get("standardization"))
    if "không đạt" in text or "khong dat" in text:
        return APPROVED_NO_STANDARDIZATION
    if "đạt" in text or "dat" in text:
        return APPROVED_STANDARDIZATION
    return APPROVED_NO_STANDARDIZATION


def resolve_status(data_row: dict[str, Any], score_row: dict[str, Any] | None) -> IdeaStatus:
    approved_score = score_row.get("approved_score") if score_row else None
    total_score = score_row.get("total_score_xn") if score_row else None
    status_text = normalize_status_text(data_row.get("status_text"))
    if approved_score and approved_score > 0:
        return IdeaStatus.APPROVED
    if "trùng lặp" in status_text or "trung lap" in status_text:
        return IdeaStatus.REJECTED
    if "chưa đủ thông tin" in status_text or "chua du thong tin" in status_text:
        return IdeaStatus.REJECTED
    if "được duyệt" in status_text or "duoc duyet" in status_text or "chỉ xét xn" in status_text or "chi xet xn" in status_text:
        return IdeaStatus.DEPT_APPROVED
    if total_score and total_score > 0:
        return IdeaStatus.DEPT_APPROVED
    return IdeaStatus.SUBMITTED


def ensure_admin_user(db) -> User:
    admin = db.query(User).filter(User.employee_code == "ADMIN").first()
    if admin is None:
        raise RuntimeError("ADMIN user was not seeded")
    return admin


def ensure_historical_score_users(db) -> tuple[User, User]:
    total_hop = db.query(Unit).filter(Unit.name == "P. Tổng hợp").first()
    if total_hop is None:
        raise RuntimeError("Unit P. Tổng hợp was not seeded")

    def ensure_user(employee_code: str, full_name: str, role: str) -> User:
        user = db.query(User).filter(User.employee_code == employee_code).first()
        if user is None:
            user = User(
                employee_code=employee_code,
                full_name=full_name,
                role=role,
                unit_id=total_hop.id,
                is_active=True,
            )
            db.add(user)
            db.flush()
            return user
        changed = False
        if user.full_name != full_name:
            user.full_name = full_name
            changed = True
        if user.role != role:
            user.role = role
            changed = True
        if user.unit_id != total_hop.id:
            user.unit_id = total_hop.id
            changed = True
        if user.is_active is not True:
            user.is_active = True
            changed = True
        if changed:
            db.add(user)
            db.flush()
        return user

    dept_user = ensure_user("HIST_DEPT", "Historical Dept Score", UserRole.DEPT_MANAGER.value)
    ie_user = ensure_user("HIST_IE", "Historical IE Score", UserRole.IE_MANAGER.value)
    return dept_user, ie_user


def existing_idea_keys(db) -> set[tuple[str | None, str, str]]:
    keys: set[tuple[str | None, str, str]] = set()
    for item in db.query(Idea.submitted_at, Idea.full_name, Idea.description).all():
        submitted_at = item[0].isoformat() if item[0] else None
        keys.add((submitted_at, item[1] or "", item[2] or ""))
    return keys


def make_attachment_type(index: int) -> str:
    return "before" if index % 2 == 1 else "after"


def import_rows(file_path: Path, dry_run: bool, limit: int | None = None) -> int:
    workbook, data_ws, score_ws = load_workbook_sheets(file_path)
    data_rows = read_data_rows(data_ws)
    score_rows = read_score_rows(score_ws)
    if limit is not None:
        data_rows = data_rows[:limit]

    db = SessionLocal()
    try:
        admin = ensure_admin_user(db)
        units_by_key = {normalize_unit_key(unit.name): unit for unit in db.query(Unit).all()}
        users_by_code = {normalize_code(user.employee_code): user for user in db.query(User).all()}
        seen_existing = existing_idea_keys(db)
        summary = Counter()

        for data_row in data_rows:
            summary["data_rows"] += 1
            stt = data_row["stt"]
            score_row = score_rows.get(stt)
            status = resolve_status(data_row, score_row)
            standardization_result = resolve_standardization_result(data_row, score_row)

            source_unit = data_row["unit"] or ""
            unit_lookup_name = UNIT_ALIASES.get(source_unit, UNIT_ALIASES.get(normalize_unit_key(source_unit), source_unit))
            unit = units_by_key.get(normalize_unit_key(unit_lookup_name))
            if unit is None:
                summary["skipped_missing_unit"] += 1
                print(f"SKIP unit not found for STT {stt}: {source_unit}", file=sys.stderr)
                continue

            submitted_at = data_row["submitted_at"]
            dedupe_key = (submitted_at.isoformat() if submitted_at else None, data_row["full_name"] or "", data_row["description"] or "")
            if dedupe_key in seen_existing:
                summary["skipped_existing"] += 1
                continue

            title = data_row["title"] or fallback_title(data_row["description"], stt)
            submitter = users_by_code.get(data_row["employee_code"])
            idea = Idea(
                full_name=data_row["full_name"] or "Unknown",
                employee_code=data_row["employee_code"] or None,
                participants_json=json.dumps([]),
                phone_number=data_row["phone"],
                bo_phan=data_row["bo_phan"],
                position=data_row["position"],
                title=title,
                product_code=data_row["product_code"],
                category=resolve_category(data_row["category_text"], data_row["description"]),
                description=data_row["description"] or "",
                submitter_id=submitter.id if submitter else None,
                unit_id=unit.id,
                status=status,
                is_anonymous=False,
                eligible_register_reward=bool(score_row and score_row.get("approved_score") and score_row["approved_score"] > 0),
                bod_register_approved=bool(score_row and score_row.get("approved_score") and score_row["approved_score"] > 0),
                bod_register_approved_at=score_row["registered_at"] if score_row and score_row.get("approved_score") else None,
                bod_register_approved_by_id=admin.id if score_row and score_row.get("approved_score") else None,
                council_final_score=score_row["approved_score"] if score_row and score_row.get("approved_score") else None,
                council_final_scored_at=score_row["registered_at"] if score_row and score_row.get("approved_score") else None,
                council_final_scored_by_id=admin.id if score_row and score_row.get("approved_score") else None,
                council_final_note=score_row["note"] if score_row and score_row.get("approved_score") else None,
                council_is_featured=False,
                council_reward_multiplier=None,
                submitted_at=submitted_at,
                reviewed_at=score_row["registered_at"] if score_row else submitted_at,
                approved_at=score_row["registered_at"] if score_row and score_row.get("approved_score") else None,
                rejected_at=submitted_at if status == IdeaStatus.REJECTED else None,
                rejection_reason=data_row["status_text"] if status == IdeaStatus.REJECTED else None,
            )
            db.add(idea)
            db.flush()
            seen_existing.add(dedupe_key)

            for attachment in data_row["attachments"]:
                file_id = parse_drive_file_id(attachment["url"])
                db.add(
                    FileAttachment(
                        idea_id=idea.id,
                        original_filename=f"historical-{stt}-{attachment['index']}.url",
                        stored_filename=file_id or f"historical-{stt}-{attachment['index']}",
                        file_type="url",
                        file_size=1,
                        file_path=f"drive://{file_id}" if file_id else attachment["url"],
                        storage_provider="google_drive" if file_id else "local",
                        external_file_id=file_id,
                        external_folder_id=None,
                        external_url=attachment["url"],
                        mime_type="text/uri-list",
                        attachment_type=make_attachment_type(attachment["index"]),
                    )
                )
            summary["attachments"] += len(data_row["attachments"])

            if status == IdeaStatus.DEPT_APPROVED:
                db.add(
                    IdeaReview(
                        idea_id=idea.id,
                        reviewer_id=admin.id,
                        level=ReviewLevel.DEPT_HEAD,
                        action=ReviewAction.APPROVE,
                        comment="Imported historical department approval",
                        reviewed_at=idea.reviewed_at or idea.submitted_at,
                    )
                )
            elif status == IdeaStatus.REJECTED:
                db.add(
                    IdeaReview(
                        idea_id=idea.id,
                        reviewer_id=admin.id,
                        level=ReviewLevel.COUNCIL,
                        action=ReviewAction.REJECT,
                        comment=data_row["status_text"] or "Imported historical rejection",
                        council_result_type="BCT_REJECTED",
                        reviewed_at=idea.reviewed_at or idea.submitted_at,
                    )
                )
            elif status == IdeaStatus.APPROVED and score_row:
                review_time = score_row["registered_at"] or idea.submitted_at
                db.add_all(
                    [
                        IdeaReview(
                            idea_id=idea.id,
                            reviewer_id=admin.id,
                            level=ReviewLevel.DEPT_HEAD,
                            action=ReviewAction.APPROVE,
                            comment="Imported historical department approval",
                            reviewed_at=review_time,
                        ),
                        IdeaReview(
                            idea_id=idea.id,
                            reviewer_id=admin.id,
                            level=ReviewLevel.COUNCIL,
                            action=ReviewAction.APPROVE,
                            comment=score_row["note"] or "Imported historical council result",
                            council_result_type=standardization_result,
                            reviewed_at=review_time,
                        ),
                        IdeaReview(
                            idea_id=idea.id,
                            reviewer_id=admin.id,
                            level=ReviewLevel.LEADERSHIP,
                            action=ReviewAction.APPROVE,
                            comment="Imported historical leadership approval",
                            reviewed_at=review_time,
                        ),
                    ]
                )

            if score_row:
                detailed = choose_detailed_score(score_row)
                if detailed is None or detailed.diff > 5:
                    summary["scores_skipped"] += 1
                else:
                    db.add(
                        IdeaScore(
                            idea_id=idea.id,
                            scorer_id=admin.id,
                            k1_type=detailed.k1_type,
                            k1_score=detailed.k1_score,
                            k1_note=None,
                            k2_type=detailed.k2_type,
                            k2_score=detailed.k2_score,
                            k2_selected_codes=json.dumps(detailed.k2_codes),
                            k2_time_frame=None,
                            k2_note=None,
                            k3_measure_type=detailed.k3_measure_type,
                            k3_option_code=detailed.k3_option_code,
                            k3_selected_codes=json.dumps(detailed.k3_codes),
                            k3_score=detailed.k3_score,
                            k3_value=None,
                            k3_note=score_row["note"],
                            total_score=score_row["total_score_xn"],
                            is_final=False,
                            scored_at=score_row["registered_at"] or idea.submitted_at,
                        )
                    )
                    summary["scores_imported"] += 1

                    if detailed.k3_measure_type != K3MeasureType.UNMEASURABLE.value:
                        before_seconds = score_row["before_seconds"]
                        after_seconds = score_row["after_seconds"]
                        quantity = score_row["quantity"]
                        if before_seconds and after_seconds is not None and quantity:
                            improvement_percent = 0.0
                            if before_seconds > 0:
                                improvement_percent = (before_seconds - after_seconds) / before_seconds
                            benefit_value = score_row["benefit_value"]
                            if benefit_value is None:
                                benefit_value = max(before_seconds - after_seconds, 0) * quantity * score_row["labor_second_price"]
                            db.add(
                                ActualBenefitEvaluation(
                                    idea_id=idea.id,
                                    evaluator_id=admin.id,
                                    before_seconds=before_seconds,
                                    after_seconds=after_seconds,
                                    improvement_percent=improvement_percent,
                                    quantity=quantity,
                                    labor_second_price=score_row["labor_second_price"],
                                    benefit_value=benefit_value,
                                    note=score_row["note"],
                                    evaluated_at=score_row["registered_at"] or idea.submitted_at,
                                )
                            )
                            summary["benefits_imported"] += 1

            summary["ideas_imported"] += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"Workbook: {file_path}")
        for key in sorted(summary):
            print(f"{key}: {summary[key]}")
        return 0
    finally:
        db.close()
        workbook.close()


def backfill_employee_codes(file_path: Path, dry_run: bool) -> int:
    workbook, data_ws, score_ws = load_workbook_sheets(file_path)
    data_rows = read_data_rows(data_ws)
    score_rows = read_score_rows(score_ws)

    db = SessionLocal()
    try:
        summary = Counter()
        for data_row in data_rows:
            score_row = score_rows.get(data_row["stt"])
            employee_code = data_row["employee_code"] or (score_row["employee_code"] if score_row else "")
            employee_code = normalize_code(employee_code)
            if not employee_code:
                continue

            submitted_at = data_row["submitted_at"]
            full_name = data_row["full_name"] or ""
            description = data_row["description"] or ""
            idea = (
                db.query(Idea)
                .filter(
                    Idea.submitted_at == submitted_at,
                    Idea.full_name == full_name,
                    Idea.description == description,
                )
                .one_or_none()
            )
            if idea is None:
                summary["not_found"] += 1
                continue
            if normalize_code(idea.employee_code):
                summary["already_has_code"] += 1
                continue

            idea.employee_code = employee_code
            db.add(idea)
            summary["updated"] += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"Workbook: {file_path}")
        for key in sorted(summary):
            print(f"{key}: {summary[key]}")
        return 0
    finally:
        db.close()
        workbook.close()


def rebuild_approved_scores(file_path: Path, dry_run: bool) -> int:
    workbook, data_ws, score_ws = load_workbook_sheets(file_path)
    data_rows = read_data_rows(data_ws)
    score_rows = read_score_rows(score_ws)

    db = SessionLocal()
    try:
        dept_user, ie_user = ensure_historical_score_users(db)
        summary = Counter()
        for data_row in data_rows:
            score_row = score_rows.get(data_row["stt"])
            if not score_row or (score_row.get("approved_score") or 0) <= 0:
                continue

            candidate = choose_detailed_score(score_row)
            if candidate is None:
                summary["skipped_no_candidate"] += 1
                continue

            idea = (
                db.query(Idea)
                .filter(
                    Idea.submitted_at == data_row["submitted_at"],
                    Idea.full_name == (data_row["full_name"] or ""),
                    Idea.description == (data_row["description"] or ""),
                )
                .one_or_none()
            )
            if idea is None:
                summary["idea_not_found"] += 1
                continue

            existing_scores = db.query(IdeaScore).filter(IdeaScore.idea_id == idea.id).all()
            for row in existing_scores:
                db.delete(row)
            summary["deleted_scores"] += len(existing_scores)

            scored_at = score_row["registered_at"] or idea.submitted_at
            score_payload = dict(
                idea_id=idea.id,
                k1_type=candidate.k1_type,
                k1_score=candidate.k1_score,
                k1_note=None,
                k2_type=candidate.k2_type,
                k2_score=candidate.k2_score,
                k2_selected_codes=json.dumps(candidate.k2_codes),
                k2_time_frame=None,
                k2_note=None,
                k3_measure_type=candidate.k3_measure_type,
                k3_option_code=candidate.k3_option_code,
                k3_selected_codes=json.dumps(candidate.k3_codes),
                k3_score=candidate.k3_score,
                k3_value=None,
                k3_note=score_row["note"],
                total_score=score_row["total_score_xn"],
                is_final=False,
                scored_at=scored_at,
            )
            db.add(IdeaScore(scorer_id=dept_user.id, **score_payload))
            db.add(IdeaScore(scorer_id=ie_user.id, **score_payload))
            summary["ideas_rebuilt"] += 1
            summary["scores_created"] += 2

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"Workbook: {file_path}")
        for key in sorted(summary):
            print(f"{key}: {summary[key]}")
        return 0
    finally:
        db.close()
        workbook.close()


def clear_historical_scores(dry_run: bool) -> int:
    db = SessionLocal()
    try:
        historical_ids_sql = "SELECT DISTINCT idea_id FROM file_attachments WHERE original_filename LIKE 'historical-%'"
        deleted_scores = db.execute(text(f"DELETE FROM idea_scores WHERE idea_id IN ({historical_ids_sql})")).rowcount
        updated_ideas = db.execute(
            text(
                f"""
                UPDATE ideas
                SET council_final_score = NULL,
                    council_final_scored_at = NULL,
                    council_final_scored_by_id = NULL,
                    council_final_note = NULL,
                    council_reward_multiplier = NULL,
                    council_is_featured = false
                WHERE id IN ({historical_ids_sql})
                  AND (
                      council_final_score IS NOT NULL
                      OR council_final_scored_at IS NOT NULL
                      OR council_final_scored_by_id IS NOT NULL
                      OR council_final_note IS NOT NULL
                      OR council_reward_multiplier IS NOT NULL
                      OR council_is_featured IS NOT FALSE
                  )
                """
            )
        ).rowcount

        hist_users = db.execute(
            text("SELECT id, employee_code FROM users WHERE employee_code IN ('HIST_DEPT', 'HIST_IE') ORDER BY employee_code")
        ).fetchall()
        deleted_users = 0
        for user_id, _employee_code in hist_users:
            ref_count = db.execute(text("SELECT COUNT(*) FROM idea_scores WHERE scorer_id = :user_id"), {"user_id": user_id}).scalar()
            if ref_count == 0:
                deleted_users += db.execute(text("DELETE FROM users WHERE id = :user_id"), {"user_id": user_id}).rowcount

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"deleted_scores: {deleted_scores}")
        print(f"updated_ideas: {updated_ideas}")
        print(f"deleted_users: {deleted_users}")
        return 0
    finally:
        db.close()


def bypass_no_score_to_library(file_path: Path, dry_run: bool) -> int:
    workbook, data_ws, score_ws = load_workbook_sheets(file_path)
    data_rows = read_data_rows(data_ws)
    score_rows = read_score_rows(score_ws)
    no_score_rows = [row for row in data_rows if row["stt"] not in score_rows]

    db = SessionLocal()
    try:
        admin = ensure_admin_user(db)
        updated = 0
        council_created = 0
        leadership_created = 0
        dept_created = 0
        not_found = 0

        for data_row in no_score_rows:
            idea = (
                db.query(Idea)
                .filter(
                    Idea.submitted_at == data_row["submitted_at"],
                    Idea.full_name == (data_row["full_name"] or ""),
                    Idea.description == (data_row["description"] or ""),
                )
                .one_or_none()
            )
            if idea is None:
                not_found += 1
                continue

            idea.status = IdeaStatus.APPROVED
            idea.approved_at = idea.approved_at or idea.submitted_at
            idea.reviewed_at = idea.reviewed_at or idea.submitted_at
            idea.rejected_at = None
            idea.rejection_reason = None
            db.add(idea)
            updated += 1

            has_dept = any(str(r.level).split(".")[-1] == "DEPT_HEAD" for r in idea.reviews)
            has_council = any(str(r.level).split(".")[-1] == "COUNCIL" for r in idea.reviews)
            has_leadership = any(str(r.level).split(".")[-1] == "LEADERSHIP" for r in idea.reviews)

            if not has_dept:
                db.add(
                    IdeaReview(
                        idea_id=idea.id,
                        reviewer_id=admin.id,
                        level=ReviewLevel.DEPT_HEAD,
                        action=ReviewAction.APPROVE,
                        comment="Historical bypass to library without score",
                        reviewed_at=idea.submitted_at,
                    )
                )
                dept_created += 1
            if not has_council:
                db.add(
                    IdeaReview(
                        idea_id=idea.id,
                        reviewer_id=admin.id,
                        level=ReviewLevel.COUNCIL,
                        action=ReviewAction.APPROVE,
                        comment="Historical bypass to non-standardization library without score",
                        council_result_type=APPROVED_NO_STANDARDIZATION,
                        reviewed_at=idea.submitted_at,
                    )
                )
                council_created += 1
            if not has_leadership:
                db.add(
                    IdeaReview(
                        idea_id=idea.id,
                        reviewer_id=admin.id,
                        level=ReviewLevel.LEADERSHIP,
                        action=ReviewAction.APPROVE,
                        comment="Historical bypass to library without score",
                        reviewed_at=idea.submitted_at,
                    )
                )
                leadership_created += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"Workbook: {file_path}")
        print(f"updated: {updated}")
        print(f"dept_created: {dept_created}")
        print(f"council_created: {council_created}")
        print(f"leadership_created: {leadership_created}")
        print(f"not_found: {not_found}")
        return 0
    finally:
        db.close()
        workbook.close()


def clear_all_idea_data(dry_run: bool) -> int:
    db = SessionLocal()
    try:
        statements = [
            ("idea_score_revisions", "DELETE FROM idea_score_revisions"),
            ("standardized_idea_replications", "DELETE FROM standardized_idea_replications"),
            ("actual_benefit_evaluations", "DELETE FROM actual_benefit_evaluations"),
            ("payment_slips", "DELETE FROM payment_slips"),
            ("idea_reviews", "DELETE FROM idea_reviews"),
            ("idea_scores", "DELETE FROM idea_scores"),
            ("file_attachments", "DELETE FROM file_attachments"),
            ("ideas", "DELETE FROM ideas"),
        ]
        deleted: dict[str, int] = {}
        for label, sql in statements:
            deleted[label] = db.execute(text(sql)).rowcount

        reset_sqls = [
            "ALTER SEQUENCE IF EXISTS idea_score_revisions_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS standardized_idea_replications_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS actual_benefit_evaluations_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS payment_slips_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS idea_reviews_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS idea_scores_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS file_attachments_id_seq RESTART WITH 1",
            "ALTER SEQUENCE IF EXISTS ideas_id_seq RESTART WITH 1",
        ]
        for sql in reset_sqls:
            db.execute(text(sql))

        if dry_run:
            db.rollback()
        else:
            db.commit()

        for label in (
            "idea_score_revisions",
            "standardized_idea_replications",
            "actual_benefit_evaluations",
            "payment_slips",
            "idea_reviews",
            "idea_scores",
            "file_attachments",
            "ideas",
        ):
            print(f"{label}: {deleted[label]}")
        return 0
    finally:
        db.close()


def import_historical_benefits(file_path: Path, dry_run: bool) -> int:
    workbook, data_ws, score_ws = load_workbook_sheets(file_path)
    data_rows = read_data_rows(data_ws)
    score_rows = read_score_rows(score_ws)

    db = SessionLocal()
    try:
        admin = ensure_admin_user(db)
        summary = Counter()

        for data_row in data_rows:
            score_row = score_rows.get(data_row["stt"])
            if not score_row:
                continue

            before_seconds = score_row.get("before_seconds")
            after_seconds = score_row.get("after_seconds")
            quantity = score_row.get("quantity")
            labor_second_price = score_row.get("labor_second_price") or 6.14
            benefit_value = score_row.get("benefit_value")

            has_measurable_inputs = before_seconds is not None and after_seconds is not None and quantity not in (None, 0)
            if benefit_value is None and not has_measurable_inputs:
                summary["skipped_missing_inputs"] += 1
                continue

            idea = (
                db.query(Idea)
                .filter(
                    Idea.submitted_at == data_row["submitted_at"],
                    Idea.full_name == (data_row["full_name"] or ""),
                    Idea.description == (data_row["description"] or ""),
                )
                .one_or_none()
            )
            if idea is None:
                summary["idea_not_found"] += 1
                continue

            normalized_before = before_seconds if before_seconds is not None else 0.0
            normalized_after = after_seconds if after_seconds is not None else 0.0
            normalized_quantity = quantity if quantity not in (None, 0) else 1

            improvement_percent = 0.0
            if before_seconds is not None and after_seconds is not None and before_seconds > 0:
                improvement_percent = (before_seconds - after_seconds) / before_seconds

            computed_benefit_value = benefit_value
            if computed_benefit_value is None:
                computed_benefit_value = max(normalized_before - normalized_after, 0) * normalized_quantity * labor_second_price

            existing = db.query(ActualBenefitEvaluation).filter(ActualBenefitEvaluation.idea_id == idea.id).one_or_none()
            if existing is None:
                existing = ActualBenefitEvaluation(
                    idea_id=idea.id,
                    evaluator_id=admin.id,
                )
                db.add(existing)
                summary["created"] += 1
            else:
                summary["updated"] += 1

            existing.evaluator_id = admin.id
            existing.before_seconds = normalized_before
            existing.after_seconds = normalized_after
            existing.improvement_percent = improvement_percent
            existing.quantity = normalized_quantity
            existing.labor_second_price = labor_second_price
            existing.benefit_value = computed_benefit_value
            existing.note = score_row.get("note")

        if dry_run:
            db.rollback()
        else:
            db.commit()

        print(f"Workbook: {file_path}")
        for key in sorted(summary):
            print(f"{key}: {summary[key]}")
        return 0
    finally:
        db.close()
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import historical Golden Idea rows from Excel")
    parser.add_argument("--file", default=str(DEFAULT_XLSX), help="Path to historical workbook")
    parser.add_argument("--dry-run", action="store_true", help="Parse and simulate inserts, then rollback")
    parser.add_argument("--limit", type=int, default=None, help="Import only the first N data rows")
    parser.add_argument(
        "--backfill-employee-code-only",
        action="store_true",
        help="Update imported historical ideas missing employee_code using data.I, then score.E as fallback",
    )
    parser.add_argument(
        "--rebuild-approved-scores",
        action="store_true",
        help="Rebuild level-1 and level-2 scores for historical approved ideas using AN + Q:AM",
    )
    parser.add_argument(
        "--clear-historical-scores",
        action="store_true",
        help="Delete historical idea_scores and clear historical council final score fields",
    )
    parser.add_argument(
        "--bypass-no-score-to-library",
        action="store_true",
        help="Bypass historical ideas that exist in data but not in Điểm into APPROVED_NO_STANDARDIZATION library",
    )
    parser.add_argument(
        "--clear-all-idea-data",
        action="store_true",
        help="Delete all idea-related data while keeping users, units, and score criteria",
    )
    parser.add_argument(
        "--import-historical-benefits",
        action="store_true",
        help="Import actual benefit data from historical Điểm sheet when measurable inputs are available",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    file_path = Path(args.file).expanduser().resolve()
    if not file_path.exists():
        print(f"Excel file not found: {file_path}", file=sys.stderr)
        return 1

    try:
        initialize_database()
        if args.backfill_employee_code_only:
            return backfill_employee_codes(file_path=file_path, dry_run=args.dry_run)
        if args.rebuild_approved_scores:
            return rebuild_approved_scores(file_path=file_path, dry_run=args.dry_run)
        if args.clear_historical_scores:
            return clear_historical_scores(dry_run=args.dry_run)
        if args.bypass_no_score_to_library:
            return bypass_no_score_to_library(file_path=file_path, dry_run=args.dry_run)
        if args.clear_all_idea_data:
            return clear_all_idea_data(dry_run=args.dry_run)
        if args.import_historical_benefits:
            return import_historical_benefits(file_path=file_path, dry_run=args.dry_run)
        return import_rows(file_path=file_path, dry_run=args.dry_run, limit=args.limit)
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
